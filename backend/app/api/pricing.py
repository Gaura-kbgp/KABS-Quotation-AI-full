from fastapi import APIRouter, UploadFile, File, BackgroundTasks
from app.core.supabase_client import get_supabase
from app.utils.cabinet_utils import (
    compress_sku, detect_category, parse_sku_dimensions,
    find_nearest_cabinet_match, classify_cabinet_type,
    normalize_collection_name, strip_catalog_suffix, strip_drawing_suffix,
    DRAWING_SUFFIX_TO_CATALOG,
)
from app.utils.excel_processor import parse_specifications_python
from app.utils.pdf_processor import parse_pricing_pdf
from thefuzz import process, fuzz
import datetime
import uuid
import os
import shutil
import re
import traceback
import json
from concurrent.futures import ThreadPoolExecutor

_CONFIG_CACHE = {}
_MFG_DB_CACHE = {'sku': {}, 'col': {}}
_TIER_KEYS  = ["PRIME", "PREMIUM", "ELITE", "CHOICE", "SELECT", "VALUE", "STANDARD"]
_WOOD_KEYS  = ["CHERRY", "MAPLE", "PAINTED", "DURAFORM", "OAK", "ASH", "HICKORY",
               "BIRCH", "ALDER", "WALNUT", "WHITE"]

WELLBORN_STRUCTURE = {
    # Col B — CHERRY and DURAFORM variants only (no MAPLE)
    "ELITE CHERRY / ELITE DURAFORM (TEXTURED)": [
        "CANYON CHERRY", "CANYON DFO CHERRY", "CANYON DFO DURAFORM (TEXTURED)",
        "DURANGO CHERRY", "ELDERIDGE CHERRY",
    ],
    # Col C — PREMIUM CHERRY + ELITE MAPLE + ELITE PAINTED (source: Excel row 2 col C)
    "PREMIUM CHERRY / PREMIUM DURAFORM (TEXTURED) / ELITE MAPLE / ELITE PAINTED": [
        "ABILENE CHERRY", "ABILENE DFO CHERRY",
        "BELCOURT CHERRY", "BELCOURT DFO CHERRY",
        "CANYON MAPLE", "CANYON DFO MAPLE", "CANYON PAINTED", "CANYON DFO PAINTED",
        "CLAYTON CHERRY", "CLAYTON DFO CHERRY",
        "DURANGO MAPLE", "DURANGO PAINTED",
        "ELDERIDGE MAPLE", "ELDERIDGE PAINTED",
        "LUBBOCK DURAFORM (TEXTURED)", "LUBBOCK CHERRY",
    ],
    # Col D — PREMIUM MAPLE, PAINTED, DURAFORM (source: Excel row 2 col D)
    "PRIME CHERRY / PREMIUM MAPLE / PREMIUM PAINTED / PREMIUM DURAFORM (NON-TEXTURED)": [
        "ABILENE DURAFORM", "ABILENE DFO DURAFORM",
        "ABILENE MAPLE", "ABILENE DFO MAPLE", "ABILENE PAINTED", "ABILENE DFO PAINTED",
        "BELCOURT MAPLE", "BELCOURT DFO MAPLE", "BELCOURT PAINTED", "BELCOURT DFO PAINTED",
        "CLAYTON MAPLE", "CLAYTON DFO MAPLE", "CLAYTON PAINTED", "CLAYTON DFO PAINTED",
        "DENVER CHERRY", "DENVER DFO CHERRY",
        "LUBBOCK DURAFORM (NON-TEXTURED)", "LUBBOCK MAPLE", "LUBBOCK PAINTED",
        "OXRIDGE MAPLE", "OXRIDGE DFO MAPLE", "OXRIDGE PAINTED", "OXRIDGE DFO PAINTED",
    ],
    # Col E (source: Excel row 2 col E)
    "PRIME MAPLE / PRIME PAINTED / PRIME DURAFORM": [
        "BANDERA MAPLE", "BANDERA DFO MAPLE", "BANDERA PAINTED", "BANDERA DFO PAINTED",
        "COOPER MAPLE", "COOPER DFO MAPLE", "COOPER PAINTED", "COOPER DFO PAINTED",
        "DENVER DURAFORM", "DENVER DFO DURAFORM",
        "DENVER MAPLE", "DENVER DFO MAPLE", "DENVER PAINTED", "DENVER DFO PAINTED",
    ],
    # Col F (source: Excel row 2 col F; CARSON DFO DUAFORM is a typo in the spreadsheet)
    "CHOICE DURAFORM / CHOICE MAPLE / CHOICE PAINTED": [
        "BARREN DURAFORM", "BARREN MAPLE", "BARREN PAINTED",
        "CARSON DURAFORM", "CARSON DFO DURAFORM",
    ],
    # Col G
    "BASE": ["BOERNE HARDWOOD"],
}

# NEW: Global Cache for fully built lookup maps
# key: manufacturer_id, value: { 'maps': lookup_maps, 'timestamp': time.time() }
GLOBAL_LOOKUP_CACHE = {}

# In-memory job tracker for async spec-book uploads
# key: job_id, value: { status, progress, message, count, fileName, error }
SPEC_UPLOAD_JOBS: dict = {}

def _fetch_pricing_data_internal(supabase, manufacturer_id: str, project_skus: set, project_collections: set = None) -> list:
    """Consolidated fetcher that handles both full load, targeted SKU, and collection-based fetch."""
    page_size = 1000
    pricing_data = []
    seen_ids: set = set()

    # 1. Total count
    try:
        count_res = supabase.table("manufacturer_pricing").select("id", count="exact").eq("manufacturer_id", manufacturer_id).execute()
        total = count_res.count or 0
    except Exception as e:
        print(f"ERROR fetching count: {e}")
        return []
    
    print(f"DEBUG: Catalog total items: {total}")

    # Use targeted/collection fetch if catalog is large (>5000)
    use_targeted = total > 5000
    
    if not use_targeted:
        # Full load for small catalogs (Parallel)
        def fetch_chunk(start, end):
            try:
                res = supabase.table("manufacturer_pricing").select("id,sku,price,collection_name,door_style") \
                    .eq("manufacturer_id", manufacturer_id).range(start, end).execute()
                return res.data or []
            except Exception as e:
                print(f"ERROR fetching range {start}-{end}: {e}")
                return []

        ranges = [(i, i + page_size - 1) for i in range(0, total, page_size)]
        with ThreadPoolExecutor(max_workers=10) as executor:
            batch_results = list(executor.map(lambda r: fetch_chunk(*r), ranges))
        
        for batch in batch_results:
            for row in batch:
                rid = row.get('id')
                if rid and rid not in seen_ids:
                    seen_ids.add(rid)
                    pricing_data.append(row)
    else:
        # Targeted SKU + Collection fetch
        sku_list = list(project_skus) if project_skus else []
        col_list = list(project_collections) if project_collections else []
        
        # A. Fetch by Collection (Broad match)
        if col_list:
            print(f"DEBUG: Fetching by collections: {col_list}")
            for cname in col_list:
                if not cname or len(cname) < 2: continue
                try:
                    # 1. Get total count for this specific collection to handle pagination
                    c_count_res = supabase.table("manufacturer_pricing").select("id", count="exact") \
                        .eq("manufacturer_id", manufacturer_id) \
                        .ilike("collection_name", f"%{cname}%").execute()
                    c_total = c_count_res.count or 0
                    
                    if c_total > 0:
                        # 2. Paginate fetch (Supabase defaults to 1000 limit)
                        for start in range(0, c_total, 1000):
                            end = start + 999
                            res = supabase.table("manufacturer_pricing").select("id,sku,price,collection_name,door_style") \
                                .eq("manufacturer_id", manufacturer_id) \
                                .ilike("collection_name", f"%{cname}%") \
                                .range(start, end).execute()
                            for row in (res.data or []):
                                rid = row.get('id')
                                if rid and rid not in seen_ids:
                                    seen_ids.add(rid); pricing_data.append(row)
                except Exception as e:
                    print(f"ERROR fetching collection {cname}: {e}")

        # B. Targeted SKU fetch (Parallel)
        if sku_list:
            print(f"DEBUG: Using targeted parallel fetch for {len(sku_list)} SKUs")
            def fetch_skus(chunk):
                try:
                    res = supabase.table("manufacturer_pricing").select("id,sku,price,collection_name,door_style") \
                        .eq("manufacturer_id", manufacturer_id).in_("sku", chunk).execute()
                    return res.data or []
                except Exception as e:
                    print(f"ERROR fetching SKU chunk: {e}")
                    return []

            chunks = [sku_list[i : i + 200] for i in range(0, len(sku_list), 200)]
            with ThreadPoolExecutor(max_workers=10) as executor:
                batch_results = list(executor.map(fetch_skus, chunks))
            
            for batch in batch_results:
                for row in batch:
                    rid = row.get('id')
                    if rid and rid not in seen_ids:
                        seen_ids.add(rid); pricing_data.append(row)
    
    return pricing_data

def get_manufacturer_pricing_maps(supabase, manufacturer_id: str, _manufacturer_hint: str = "", project_skus: set = None, project_collections: set = None):
    """
    Returns built lookup maps for a manufacturer, using a global memory cache if available.
    """
    now = datetime.datetime.now().timestamp()
    
    # Use cache if available (Full catalog cache covers any targeted request)
    if manufacturer_id in GLOBAL_LOOKUP_CACHE:
        entry = GLOBAL_LOOKUP_CACHE[manufacturer_id]
        if now - entry['timestamp'] < 900: # 15 mins
            print(f"DEBUG: Using CACHED full lookup maps for {manufacturer_id}")
            return entry['maps']

    # 2. Fetch all rows (or targeted rows) using consolidated fetcher
    pricing_data = _fetch_pricing_data_internal(supabase, manufacturer_id, project_skus or set(), project_collections or set())

    # 3. Build maps
    lookup_maps = _build_lookup_maps(pricing_data)

    # 4. Cache ONLY if it's a full load (no project_skus) and has data
    if not project_skus and lookup_maps.get('all_catalog_rows'):
        GLOBAL_LOOKUP_CACHE[manufacturer_id] = {
            'maps': lookup_maps,
            'timestamp': now
        }
    return lookup_maps

def _build_lookup_maps(pricing_data: list):
    from app.utils.cabinet_utils import (
        detect_category, strip_catalog_suffix, compress_sku, 
        normalize_collection_name, parse_sku_dimensions, get_cabinet_section,
        classify_cabinet_type
    )
    
    lookup_maps = {
        'local': {}, 'global': {}, 'compressed': {}, 'dim': {},
        'col_skus': {}, 'category_items': {}, 'category_sums': {},
        'all_catalog_rows': [],
        # NEW: SECTION-INDEXED rows for 10x faster dimension matching
        'section_rows': {},
        'labor': {} # NEW: { sku: points }
    }

    # MEMOIZATION: Drastically speed up map building for multi-collection catalogs
    _sku_info_cache = {}
    _norm_col_cache = {}

    for p in pricing_data:
        sku = str(p['sku']).strip().upper()
        price = float(p.get('price') or 0)
        col = str(p.get('collection_name', '')).strip().upper()
        style = str(p.get('door_style', '')).strip().upper()
        
        # 1. Memoized SKU properties (independent of collection)
        if sku not in _sku_info_cache:
            sku_base = strip_catalog_suffix(sku)
            _sku_info_cache[sku] = {
                "base": sku_base,
                "ct": classify_cabinet_type(sku_base),
                "cs": get_cabinet_section(sku_base),
                "dims": parse_sku_dimensions(sku_base),
                "ac": re.sub(r'[^A-Z0-9]', '', sku_base.upper()),
                "comp": compress_sku(sku),
                "cat": detect_category(sku)
            }
        
        s_info = _sku_info_cache[sku]
        sku_base = s_info["base"]
        ct = s_info["ct"]
        cs = s_info["cs"]
        dims_p = s_info["dims"]
        w_val = dims_p.get('width')
        h_val = dims_p.get('height')
        ac = s_info["ac"]
        comp = s_info["comp"]
        cat = s_info["cat"]

        # 2. Memoized Collection normalization
        if col not in _norm_col_cache:
            _norm_col_cache[col] = normalize_collection_name(col)
        norm_col = _norm_col_cache[col]
        
        # ── Labor Points Extraction (Wellborn Accessory Sheets) ──
        if col == "LABOR_POINTS":
            lookup_maps['labor'][sku] = price
            continue
            
        # SELF-HEALING: Prevent labor point factors (0.1, 0.3, 1.0) from being used as material prices.
        # This occurs when points were mistakenly uploaded as prices in the primary collection.
        from app.utils.cabinet_utils import detect_category
        item_cat = detect_category(sku)
        # Aggressive filter for any item with 'point-like' prices in labor-prone categories
        is_labor_prone = item_cat in ("Molding & Trim", "Universal Fillers", "Accessories", "Hardwares", "Panels")
        
        if is_labor_prone and 0 < price < 5.0:
            if sku not in lookup_maps['labor']:
                lookup_maps['labor'][sku] = price
            # Skip from primary list-price maps to allow finding the real price row
            continue
        
        item = {
            "sku": sku, "price": price, "collection_name": col, "door_style": style,
            "ct": ct, "cs": cs, "w": w_val, "h": h_val,
            "ncol": norm_col, "csku": sku_base,
            "ac": ac
        }
        
        lookup_maps['all_catalog_rows'].append(item)
        
        # ── Section Indexing for Fast Nearest Match ──
        if cs:
            if cs not in lookup_maps['section_rows']:
                lookup_maps['section_rows'][cs] = []
            lookup_maps['section_rows'][cs].append(item)
        
        sku_comp_base = compress_sku(sku_base)

        # Indexing for local_map[sku|col] and local_map[sku|col|style]
        # Priority mapping: If multiple rows exist (e.g. from bad uploads), prioritize higher prices
        def set_local_priority(key, val):
            if key not in lookup_maps['local'] or val['price'] > lookup_maps['local'][key]['price']:
                lookup_maps['local'][key] = val

        if col and style: set_local_priority(f"{sku}|{col}|{style}", item)
        if col: set_local_priority(f"{sku}|{col}", item)
        
        if col:
            if col not in lookup_maps['col_skus']: lookup_maps['col_skus'][col] = []
            lookup_maps['col_skus'][col].append(sku)

        norm_col = normalize_collection_name(col)
        if norm_col and norm_col != col:
            if norm_col not in lookup_maps['col_skus']:
                lookup_maps['col_skus'][norm_col] = []
            lookup_maps['col_skus'][norm_col].append(sku)
            if col and style: set_local_priority(f"{sku}|{norm_col}|{style}", item)
            if col: set_local_priority(f"{sku}|{norm_col}", item)

        # Indexing under all component Tier + Wood combinations
        tiers_in_col = [t for t in _TIER_KEYS if t in col]
        woods_in_col = [w for w in _WOOD_KEYS if w in col]
        for t in tiers_in_col:
            for w in woods_in_col:
                variant_col = f"{t} {w}"
                if variant_col != col:
                    if variant_col not in lookup_maps['col_skus']:
                        lookup_maps['col_skus'][variant_col] = []
                    lookup_maps['col_skus'][variant_col].append(sku)
                    set_local_priority(f"{sku}|{variant_col}", item)
                    if style: set_local_priority(f"{sku}|{variant_col}|{style}", item)
        
        # Wellborn Structured Collection Indexing
        for full_col in WELLBORN_STRUCTURE.keys():
            if col and (col in full_col or full_col in col):
                set_local_priority(f"{sku}|{full_col}", item)
                if comp != sku: set_local_priority(f"{comp}|{full_col}", item)
                if style: 
                    set_local_priority(f"{sku}|{full_col}|{style}", item)
                    if comp != sku: set_local_priority(f"{comp}|{full_col}|{style}", item)
                    if sku_base != sku: set_local_priority(f"{sku_base}|{full_col}|{style}", item)
                if sku_base != sku: set_local_priority(f"{sku_base}|{full_col}", item)

                if full_col not in lookup_maps['col_skus']:
                    lookup_maps['col_skus'][full_col] = []
                lookup_maps['col_skus'][full_col].append(sku)
                if comp != sku: lookup_maps['col_skus'][full_col].append(comp)

                norm_full = normalize_collection_name(full_col)
                if norm_full != full_col:
                    set_local_priority(f"{sku}|{norm_full}", item)
                    if comp != sku: set_local_priority(f"{comp}|{norm_full}", item)
                    if norm_full not in lookup_maps['col_skus']:
                        lookup_maps['col_skus'][norm_full] = []
                    lookup_maps['col_skus'][norm_full].append(sku)

        if style: set_local_priority(f"{sku}|{style}", item)

        # Priority mapping for global_map[sku]
        if sku not in lookup_maps['global'] or price > lookup_maps['global'][sku]['price']:
            lookup_maps['global'][sku] = item
        
        if comp not in lookup_maps['compressed']: lookup_maps['compressed'][comp] = item

        # Index compressed SKU in local map
        if comp != sku:
            if col:
                set_local_priority(f"{comp}|{col}", item)
                if style: set_local_priority(f"{comp}|{col}|{style}", item)
            if norm_col and norm_col != col:
                set_local_priority(f"{comp}|{norm_col}", item)
            for t in tiers_in_col:
                for w in woods_in_col:
                    variant_col = f"{t} {w}"
                    if variant_col != col:
                        set_local_priority(f"{comp}|{variant_col}", item)
                        if style: set_local_priority(f"{comp}|{variant_col}|{style}", item)

        if sku_base and sku_base != sku:
            if col and style: set_local_priority(f"{sku_base}|{col}|{style}", item)
            if col: set_local_priority(f"{sku_base}|{col}", item)
            if style: set_local_priority(f"{sku_base}|{style}", item)
            
            if sku_base not in lookup_maps['global'] or price > lookup_maps['global'][sku_base]['price']:
                lookup_maps['global'][sku_base] = item
                
            if sku_comp_base not in lookup_maps['compressed']:
                lookup_maps['compressed'][sku_comp_base] = item
            if norm_col and norm_col != col:
                set_local_priority(f"{sku_base}|{norm_col}", item)

        stripped = re.sub(r'[\s-]*(BUTT|H|L|R|FL|S|D)$', '', sku).strip()
        if stripped != sku:
            lookup_maps['local'][f"{stripped}|{col}|{style}"] = item
            lookup_maps['local'][f"{stripped}|{col}"] = item
            lookup_maps['local'][f"{stripped}|{style}"] = item
            if stripped not in lookup_maps['global']:
                lookup_maps['global'][stripped] = item

        if cat not in lookup_maps['category_items']: lookup_maps['category_items'][cat] = []
        lookup_maps['category_items'][cat].append(item)
        if cat not in lookup_maps['category_sums']: lookup_maps['category_sums'][cat] = [0.0, 0]
        # Outlier protection: skip insane prices (likely phone numbers or parsing errors)
        if price < 50000:
            lookup_maps['category_sums'][cat][0] += price
            lookup_maps['category_sums'][cat][1] += 1

        dims = parse_sku_dimensions(sku_base if sku_base else sku)
        if dims.get('prefix') and dims.get('width'):
            dim_key = f"{dims['prefix']}|{dims['width']}|{dims.get('height') or ''}"
            if f"{dim_key}|{col}" not in lookup_maps['dim']:
                lookup_maps['dim'][f"{dim_key}|{col}"] = item
            if dim_key not in lookup_maps['dim']:
                lookup_maps['dim'][dim_key] = item
    return lookup_maps

def get_cache_path(mfg_id: str) -> str:
    path = os.path.join(os.path.dirname(__file__), "..", "..", f"mfg_config_{mfg_id}.json")
    return os.path.abspath(path)

router = APIRouter()

def find_best_match(item_code: str, room_collection: str, room_door_style: str, lookup_maps: dict, manufacturer_hint: str = ""):
    target = str(item_code or "").strip().upper()
    if not target: return None, None
    
    col = str(room_collection or "").strip().upper()
    style = str(room_door_style or "").strip().upper()
    category = detect_category(target)
    
    local_map = lookup_maps.get('local', {})
    global_map = lookup_maps.get('global', {})
    compressed_map = lookup_maps.get('compressed', {})


    def log_match(message):
        # Disabled file I/O for 10x speed improvement
        # print(f"DEBUG: {message}")
        pass

    def try_match(sku_variant: str, match_type_suffix: str, include_global: bool = True, override_col: str = None):
        if not sku_variant: return None
        target_col = override_col if override_col else col
        # 1. Strict SKU + Col + Style
        key1 = f"{sku_variant}|{target_col}|{style}"
        if key1 in local_map:
            log_match(f"MATCH: {key1} (Local Spec)")
            return local_map[key1], f"EXACT_SPEC_{match_type_suffix}"
            
        # 2. Strict SKU + Style
        key2 = f"{sku_variant}|{style}"
        if key2 in local_map:
            log_match(f"MATCH: {key2} (Local Style)")
            return local_map[key2], f"EXACT_STYLE_{match_type_suffix}"
            
        # 3. Strict SKU + Col
        key3 = f"{sku_variant}|{target_col}"
        if key3 in local_map:
            log_match(f"MATCH: {key3} (Local Col)")
            return local_map[key3], f"EXACT_COL_{match_type_suffix}"
            
        # 4. Global SKU (Now optional)
        if include_global and sku_variant in global_map:
            log_match(f"MATCH: {sku_variant} (Global)")
            return global_map[sku_variant], f"EXACT_GLOBAL_{match_type_suffix}"
        return None

    log_match(f"PROBING: {target} (Room: {col} / {style})")

    # PRE-FLIGHT: Gather all variants to probe
    # We will try LOCAL match for all variants first, then GLOBAL as fallback.
    clean_target = re.sub(r'[\(\{\[].*?[\)\}\]]', '', target).strip()
    clean_target = re.sub(r'\s*(-EST|EST\.)$', '', clean_target).strip()
    
    # Suffix Bridge: convert drawing suffixes to catalog suffixes.
    # Handles both spaced ("B30 BUTT" → "B30BD") and concatenated ("W3624BUTT" → "W3624BD").
    bridged = None
    suffix_m = re.search(r'\s+(BUTT DOOR|BUTT|BD|LH|RH|L|R)$', clean_target)
    if suffix_m:
        suffix_token = suffix_m.group(1).strip()
        base_sku = clean_target[:suffix_m.start()].strip()
        mapped = DRAWING_SUFFIX_TO_CATALOG.get(suffix_token)
        if mapped:
            bridged = base_sku + mapped
    else:
        # Concatenated form: "W3624BUTT" → "W3624BD", "B30BD" → already BD
        conc_m = re.match(r'^([A-Z]+\d+)(BUTT DOOR|BUTT|LH|RH)$', clean_target)
        if conc_m:
            mapped = DRAWING_SUFFIX_TO_CATALOG.get(conc_m.group(2).strip())
            if mapped:
                bridged = conc_m.group(1) + mapped

    stripped_drawing = strip_drawing_suffix(clean_target)
    no_suffix_target = re.sub(r'\s*(BUTT|H|L|R|FL|S|D)$', '', clean_target).strip()
    
    _MFG_SUFFIX_RE = re.compile(r'(VENTBOX|VENT|BOX|MDBD|BLD|DWR|DW|BD|SD|MD|AS|A|EST|EST\.)$')
    mfg_stripped = _MFG_SUFFIX_RE.sub('', clean_target).strip()
    
    variants = [
        (target, "ORIGINAL"),
        (bridged, "BRIDGED") if bridged else (None, ""),
        (clean_target, "CLEANED") if clean_target != target else (None, ""),
        (stripped_drawing, "STRIPPED") if stripped_drawing != clean_target else (None, ""),
        (no_suffix_target, "NO_SUFFIX") if no_suffix_target != clean_target else (None, ""),
        (mfg_stripped, "MFG_STRIPPED") if mfg_stripped != clean_target else (None, ""),
    ]
    
    # TIER 1-3: Try LOCAL match for all variants
    norm_col = normalize_collection_name(col)
    for var, label in variants:
        if not var: continue
        # Try raw collection first
        res = try_match(var, label, include_global=False)
        if res: return res
        # Try normalized collection second
        if norm_col and norm_col != col:
            res = try_match(var, f"{label}_NORMCOL", include_global=False, override_col=norm_col)
            if res: return res

    # NEW: Try SPACED variants specifically (e.g. W3624BUTT -> W3624 BUTT)
    if "BUTT" in clean_target and " " not in clean_target:
        spaced = clean_target.replace("BUTT", " BUTT")
        res = try_match(spaced, "SPACED", include_global=False)
        if res: return res

    # NEW: Try SBN variant for Sink Bases
    if clean_target.startswith("SB") and not clean_target.startswith("SBN"):
        sbn_variant = clean_target.replace("SB", "SBN", 1)
        res = try_match(sbn_variant, "SBN_PROBE", include_global=False)
        if res: return res

    # TIER 4: Compressed match (handles W3624BUTT vs W3624 BUTT)
    comp_target = compress_sku(clean_target)
    if comp_target in compressed_map:
        return compressed_map[comp_target], "COMPRESSED"
    
    comp_stripped = compress_sku(stripped_drawing)
    if comp_stripped and comp_stripped != comp_target and comp_stripped in compressed_map:
        return compressed_map[comp_stripped], "COMPRESSED_STRIPPED"

    # TIER 4.4: SKU-exact pre-aliases for items with non-standard formats
    _NO_DIGIT_ALIASES: dict = {
        'TOUCHUPSPRAY':  ['TOUCHUP SPRAY', 'TOUCH-UP SPRAY', 'SPRAY', 'TOUCHUPSPRAY', 'TOUCHUP',
                          'TOUCHUP KIT', 'TOUCH UP SPRAY', 'TUSP', 'TS', 'SPRAY KIT',
                          'TOUCHUP SPRAY KIT', 'SPRAY TOUCH', 'TOUCH UP KIT'],
        'TOUCHUPKIT':    ['TOUCHUP KIT', 'TOUCH-UP KIT', 'TOUCHUPKIT', 'TOUCHUP', 'TOUCH UP KIT'],
        'TOUCHUP':       ['TOUCHUP KIT', 'TOUCHUP SPRAY', 'TOUCHUP', 'TOUCH-UP KIT'],
        'LIGHTRAIL':     ['LIGHT RAIL', 'LR', 'LIGHTRAIL'],
        'SHELFADJ48':    ['SHELF ADJ 48', 'SHELFADJ48', 'SHELF ADJ', 'SHELF 48',
                          'ADJUSTABLE SHELF 48', 'ADJ SHELF 48', 'ADJSHELF48',
                          'SHELF-ADJ-48', 'SHELF ADJUST 48', 'ADJSHELF', 'SHELF'],
        'SHELFADJ36':    ['SHELF ADJ 36', 'SHELFADJ36', 'SHELF 36', 'SHELF ADJ', 'ADJSHELF36', 'SHELF'],
        'SHELFADJ30':    ['SHELF ADJ 30', 'SHELFADJ30', 'SHELF 30', 'SHELF ADJ', 'ADJSHELF30', 'SHELF'],
        # BACKB48 = 48" Finished Back Panel / vent chase back board.
        # Catalog may list this dimensionlessly (FBP), by sheet (GP4896), or as an end panel.
        'BACKB48':       ['FBP48', 'FBP', 'BACK-B48', 'BACK-B', 'BACK PANEL 48', 'BACK PANEL',
                          'BACKPANEL48', 'BACKPANEL', 'FINISHED BACK PANEL', 'FINBACK48', 'FINBACK',
                          'GP4896', 'GP48', 'PANEL48', 'BACKER48', 'BACKER',
                          'BACKBOARD48', 'BOXBACK48', 'BOXBACK', 'VENTBOX48', 'VENT BACK 48',
                          'WTEP48', 'TEP48', 'EP48', 'SADW48', 'SAD48',
                          'WTEP', 'TEP', 'EP', 'SADW', 'SAD'],
        'BACKB36':       ['FBP36', 'FBP', 'BACK-B36', 'BACK PANEL 36', 'BACK PANEL',
                          'BACKPANEL36', 'BACKPANEL', 'GP3696', 'GP36', 'PANEL36',
                          'WTEP36', 'TEP36', 'EP36', 'SADW36', 'WTEP', 'TEP', 'EP'],
        'BACKB24':       ['FBP24', 'FBP', 'BACK-B24', 'BACK PANEL 24', 'BACK PANEL',
                          'BACKPANEL24', 'BACKPANEL', 'GP2496', 'GP24', 'PANEL24',
                          'WTEP24', 'TEP24', 'EP24', 'SADW24', 'WTEP', 'TEP', 'EP'],
    }
    # TIER 4.4 keyword partial-scan table — used if exact probes above all miss
    # Back panels: scan for any end-panel or generic-panel type in catalog (WTEP84 is confirmed present)
    _NO_DIGIT_SCAN: dict = {
        'TOUCHUPSPRAY':  ['TOUCH', 'SPRAY', 'TOUCHUP'],
        'TOUCHUPKIT':    ['TOUCHUP', 'TOUCH', 'KIT'],
        'TOUCHUP':       ['TOUCHUP', 'TOUCH'],
        'SHELFADJ48':    ['SHELF', 'SHELFADJ', 'ADJ'],
        'SHELFADJ36':    ['SHELF', 'SHELFADJ', 'ADJ'],
        'SHELFADJ30':    ['SHELF', 'SHELFADJ', 'ADJ'],
        'LIGHTRAIL':     ['RAIL', 'LIGHT', 'LR'],
        'BACKB48':       ['FBP', 'BACKPANEL', 'BACKER', 'BACK', 'WTEP', 'TEP', 'SADW'],
        'BACKB36':       ['FBP', 'BACKPANEL', 'BACKER', 'BACK', 'WTEP', 'TEP', 'SADW'],
        'BACKB24':       ['FBP', 'BACKPANEL', 'BACKER', 'BACK', 'WTEP', 'TEP', 'SADW'],
    }
    # SKU prefixes that should return MANUAL ($0) rather than a misleading category average
    # when no catalog match is found at all
    _MANUAL_FALLBACK_PREFIXES = ('BACKB', 'BACKF', 'FBP', 'BOXBACK')
    _is_manual_fallback = any(clean_target.startswith(p) for p in _MANUAL_FALLBACK_PREFIXES)
    if clean_target in _NO_DIGIT_ALIASES:
        for probe in _NO_DIGIT_ALIASES[clean_target]:
            res = try_match(probe, 'DIRECT_ALIAS', include_global=True)
            if res:
                return res
        # TIER 4.45: Partial catalog key scan for these hard-to-match accessories
        scan_kws = _NO_DIGIT_SCAN.get(clean_target, [])
        _scan_col_best = None
        _scan_global_best = None
        for gsku, gitem in global_map.items():
            gsku_up = gsku.upper()
            if gsku_up == clean_target:
                continue
            for kw in scan_kws:
                if kw in gsku_up:
                    gcol = str(gitem.get('collection_name') or '').upper()
                    if gcol == col and _scan_col_best is None:
                        _scan_col_best = (gitem, f"PARTIAL_SCAN_{clean_target}_COL")
                    elif _scan_global_best is None:
                        _scan_global_best = (gitem, f"PARTIAL_SCAN_{clean_target}_GLOBAL")
                    break
        if _scan_col_best:
            return _scan_col_best
        if _scan_global_best:
            return _scan_global_best
        # Nothing in catalog at all — for back panels return MANUAL ($0) rather than a
        # misleading hardware-category average (e.g. $16 avg of hinges/clips).
        if _is_manual_fallback:
            return {
                "sku": f"{clean_target}_MANUAL",
                "price": 0.0,
                "collection_name": col or "Unknown",
                "door_style": style or "Unknown",
                "price_ref": (
                    f"Back panel '{clean_target}' not found in catalog — "
                    f"enter the manufacturer list price per panel manually."
                )
            }, "MANUAL_PRICING_REQUIRED"

    # TIER 4.5: Molding/Alias Probing (Local and Global)
    _MOLDING_ALIASES = {
        'BTK': ['BTK', 'TK', 'BK'],      # Base Toe Kick
        'TK':  ['TK', 'BTK'],
        'FL':  ['FL', 'LR', 'FLR'],      # Filler Light Rail
        'LR':  ['LR', 'FL', 'FLR'],      # Light Rail
        'OCM': ['OCM', 'OC', 'CM', 'OCORNER', 'OUTSIDE'], # Outside Corner Molding
        'OC':  ['OC', 'OCM', 'OCORNER'],
        'ICM': ['ICM', 'IC', 'ICORNER'],  # Inside Corner Molding
        'IC':  ['IC', 'ICM'],
        'SCM': ['SCM', 'SM', 'SCRIBE', 'SC'],  # Scribe Molding
        'SM':  ['SM', 'SCM', 'SHOE'],
        'SHM': ['SHM', 'SHAKER'],
        'CM':  ['CM', 'OCM', 'CROWN', 'CRN'],   # Crown Molding
        'QM':  ['QM', 'QR', 'QUARTER'],          # Quarter Round
        'RR':  ['RR', 'LR'],             # Return Rail
        'REF': ['REP', 'REFP', 'REF'],   # Refrigerator End Panel
        'REP': ['REF', 'REFP', 'REP'],
        'BACKB': ['BACKB', 'BACK-B', 'FBP', 'BACKPANEL', 'BOXBACK', 'BACKER', 'FINBACK', 'GP'],  # Back panel / vent chase back
        'BACKF': ['BACKF', 'BACK-F', 'FBP', 'FINBACK', 'BACKPANEL'],                             # Finished back panel
        'WTEP':  ['WTEP', 'TEP', 'EP'],                                                           # Wall/toe-end panel
    }
    _CATALOG_VARIANTS = ['BLD', 'BD', 'SD', 'MD', 'MDBD', 'AS', 'A', 'S', '']
    _alias_prefix_m = re.match(r'^([A-Z]+)(\d+.*)$', clean_target)
    if _alias_prefix_m:
        ap, adigits = _alias_prefix_m.group(1), _alias_prefix_m.group(2)
        adigits_clean = _MFG_SUFFIX_RE.sub('', adigits).strip()
        if ap in _MOLDING_ALIASES:
            for alias in _MOLDING_ALIASES[ap]:
                for variant in _CATALOG_VARIANTS:
                    probe = f"{alias}{adigits_clean}{variant}"
                    if probe != clean_target:
                        # Enable include_global=True so moldings in UNIVERSAL are found
                        res = try_match(probe, f"ALIAS_{alias}", include_global=True)
                        if res: return res

    # TIER 5: GLOBAL FALLBACK for exact/bridged/spaced variants
    for var, label in variants:
        if not var: continue
        if var in global_map:
            return global_map[var], f"EXACT_GLOBAL_{label}"
    
    if "BUTT" in clean_target and " " not in clean_target:
        spaced = clean_target.replace("BUTT", " BUTT")
        if spaced in global_map:
            return global_map[spaced], "EXACT_GLOBAL_SPACED"

    # TIER 6: FUZZY COLLECTION-BOUND MATCHING
    # Normalize collection for matching
    norm_col = normalize_collection_name(col)
    col_candidates = []
    for catalog_col in lookup_maps.get('col_skus', {}).keys():
        cat_norm = normalize_collection_name(catalog_col)
        if norm_col == cat_norm or (norm_col and (norm_col in cat_norm or cat_norm in norm_col)):
            col_candidates.append(catalog_col)
            
    if col in lookup_maps.get('col_skus', {}) and col not in col_candidates:
        col_candidates.append(col)

    fuzzy_target = compress_sku(clean_target)

    # TIER 5.5: SUPER FUZZY (Handles extreme formatting diffs)
    # Uses token_set_ratio which ignores duplicates and word order
    for fuzzy_col_key in col_candidates:
        choices = lookup_maps['col_skus'].get(fuzzy_col_key, [])
        if not choices: continue
        best_sku, score = process.extractOne(fuzzy_target, choices, scorer=fuzz.token_set_ratio)
        if score >= 95:
            key = f"{best_sku}|{fuzzy_col_key}"
            match = local_map.get(key)
            if match: return match, f"SMART_FUZZY_{score}"

    # TIER 5.7: BASE CODE PROBE (W3624BUTT -> W3624)
    # Use a more aggressive base-code extractor
    smart_base = re.sub(r'[^A-Z0-9]', '', clean_target)
    smart_base = re.sub(r'(BUTT|BD|SD|MD|H|L|R)$', '', smart_base)
    if smart_base != clean_target:
        res = try_match(smart_base, "SMART_BASE", include_global=False)
        if res: return res
        # Also try spaced variant of smart base
        for fuzzy_col_key in col_candidates:
            choices = lookup_maps['col_skus'].get(fuzzy_col_key, [])
            if smart_base in choices:
                return local_map.get(f"{smart_base}|{fuzzy_col_key}"), "SMART_BASE_LOCAL"

    # TIER 6: GLOBAL FUZZY MATCH
    global_choices = list(global_map.keys())
    if global_choices:
        # Use WRatio which is the most advanced fuzzy matcher in thefuzz
        best_sku, score = process.extractOne(fuzzy_target, global_choices, scorer=fuzz.WRatio)
        if score >= 95:
            return global_map[best_sku], f"FUZZY_GLOBAL_{score}"

    # TIER 6.5: SMART DIMENSION IDENTITY (Handles SKU variations like W3624BT vs W3624BUTT)
    # If we find exactly ONE item in the collection with the same category and dimensions,
    # it's almost certainly the right item even if the SKU strings differ.
    target_dims = parse_sku_dimensions(clean_target)
    from app.utils.cabinet_utils import get_cabinet_section
    target_section = get_cabinet_section(clean_target)
    if target_dims.get('width') and target_section:
        dim_candidates = []
        tw, th = target_dims['width'], target_dims.get('height')
        for c_key in col_candidates:
            sec_items = lookup_maps.get('section_rows', {}).get(target_section, [])
            for item in sec_items:
                # Normalize collection check
                item_ncol = item.get('ncol') or normalize_collection_name(str(item.get('collection_name', '')))
                if item_ncol == c_key or item.get('collection_name') == c_key:
                    if item.get('w') == tw and item.get('h') == th:
                        dim_candidates.append(item)
        if len(dim_candidates) == 1:
            return dim_candidates[0], "SMART_DIM_IDENTITY"

    # TIER 7: DIMENSION MATCH (Collection-Bound)
    dims = parse_sku_dimensions(clean_target)
    if dims.get('prefix') and dims.get('width'):
        dim_key = f"{dims['prefix']}|{dims['width']}|{dims.get('height') or ''}"
        dim_map = lookup_maps.get('dim', {})
        for c_cand in col_candidates:
            if f"{dim_key}|{c_cand}" in dim_map:
                return dim_map[f"{dim_key}|{c_cand}"], "DIMENSION_COL"

    # TIER 8: INTELLIGENT NEAREST-CABINET (DIM)
    cabinet_type = classify_cabinet_type(clean_target)
    if cabinet_type and lookup_maps.get('all_catalog_rows'):
        nearest = find_nearest_cabinet_match(
            target_sku=clean_target,
            catalog_skus=lookup_maps, 
            collection_filter=col or None,
            manufacturer_hint=manufacturer_hint,
        )
        if nearest:
            from app.utils.cabinet_utils import parse_sku_dimensions as _psd
            t_dims, n_dims = _psd(clean_target), _psd(str(nearest.get('sku', '')))
            tw, nw, th, nh = t_dims.get('width') or 0, n_dims.get('width') or 0, t_dims.get('height') or 0, n_dims.get('height') or 0
            size_note = ""
            if tw and nw and tw != nw: size_note = f" (nearest {nw}\" wide)"
            elif th and nh and th != nh: size_note = f" (nearest {nh}\" tall)"
            ref_sku, ref_col = nearest.get('sku', ''), nearest.get('collection_name', '')
            nearest['price_ref'] = f"Ref: {ref_sku}{size_note} [{ref_col}]"
            
            # Use 'EXACT' if dimensions match perfectly even if SKU differs
            m_type = f"NEAREST_DIM_{cabinet_type.replace(' ', '_').upper()}"
            if tw == nw and (th == nh or not th or not nh):
                m_type = f"DIMENSION_MATCH_{cabinet_type.replace(' ', '_').upper()}"
            return nearest, m_type

    # TIER 8.5: DOOR / DRAWER — style-specific catalog probe
    # DOORS/DRAWERS counts arrive with no direct catalog SKU.  We try every plausible
    # catalog key before falling through. If nothing is found we return MANUAL ($0)
    # rather than the misleading 2 000+ item average.
    if target in ('DOORS', 'DRAWERS', 'DOOR', 'DRAWER'):
        is_drawer = 'DRAWER' in target
        door_probes = (
            ['DRAWER', 'DWR', 'DRW', 'DRAWER BOX', 'DRAWERBOX', 'DBX', 'DRW BOX',
             'DRAWER FRONT', 'DRAWERFRONT', 'DWR FRONT', 'DRAWER FACE', 'DWRFRONT']
            if is_drawer else
            ['DOOR', 'DOOR PANEL', 'DOORPANEL', 'DP', 'DR', 'REPLACEMENT DOOR', 'RPDOOR', 'RPD',
             'DOOR FRONT', 'DOORFRONT', 'DOOR FACE', 'REPL DOOR', 'REPLDOOR', 'REPDOOR',
             'OVERLAY DOOR', 'INSET DOOR', 'SLAB DOOR', 'PANEL DOOR', 'FRAME DOOR']
        )
        for probe in door_probes:
            res = try_match(probe, 'DOOR_CATALOG', include_global=True)
            if res:
                return res
        # Try the full door-style name as a catalog SKU
        # (some manufacturers list prices under the style name itself)
        if style:
            res = try_match(style, 'STYLE_AS_DOOR', include_global=True)
            if res:
                return res
            # Try every individual word of the door-style name (length > 3 to skip short words)
            style_words = [w for w in style.split() if len(w) > 3]
            for word in style_words:
                res = try_match(word, 'STYLE_WORD', include_global=True)
                if res:
                    return res
            # Try style + DOOR/DRAWER compound probes
            for door_word in (['DRAWER', 'DWR'] if is_drawer else ['DOOR', 'DP']):
                res = try_match(f"{style} {door_word}", 'STYLE_COMPOUND', include_global=True)
                if res:
                    return res
                first_word = style.split()[0] if style.split() else ''
                if first_word:
                    res = try_match(f"{first_word} {door_word}", 'STYLE_WORD_COMPOUND', include_global=True)
                    if res:
                        return res
        # TIER 8.6: Partial scan — walk global catalog keys for any SKU containing DOOR/DRAWER keywords.
        # Prefer items in the same collection; fall back to any matching catalog entry.
        # Skip exact keyword matches (e.g. sku == "DOOR") — those were already tried above.
        _scan_terms = ['DRAWER', 'DWR', 'DRAW'] if is_drawer else ['DOOR', 'DP']
        _scan_skip  = {'DOOR', 'DRAWER', 'DOORS', 'DRAWERS', 'DWR', 'DP', 'DR', 'DRW'}
        _found_global = None
        for gsku, gitem in global_map.items():
            gsku_up = gsku.upper()
            if gsku_up in _scan_skip:
                continue
            for term in _scan_terms:
                if term in gsku_up:
                    gcol = str(gitem.get('collection_name') or '').upper()
                    if gcol == col:
                        return gitem, f"PARTIAL_SCAN_LOCAL_{term}"
                    if _found_global is None:
                        _found_global = gitem  # keep first global candidate
                    break
        if _found_global is not None:
            return _found_global, "PARTIAL_SCAN_GLOBAL"

        # Nothing in catalog — return $0 so the user knows to enter the price manually
        label = 'Drawer' if is_drawer else 'Door'
        style_label = style if style else 'selected door style'
        return {
            "sku": f"{target}_MANUAL",
            "price": 0.0,
            "collection_name": col or "Unknown",
            "door_style": style or "Unknown",
            "price_ref": f"{label} price not found in catalog for {style_label} — enter unit price manually"
        }, "MANUAL_PRICING_REQUIRED"

    # TIER 9: CATEGORY FALLBACK
    cat_sums = lookup_maps.get('category_sums', {})
    if category in cat_sums:
        total_price, count = cat_sums[category]
        if count > 0:
            avg_price = total_price / count
            cat_items = lookup_maps.get('category_items', {}).get(category, [])
            ref_msg = f"Avg of {count} {category} items"
            if cat_items:
                best_item = cat_items[0]
                ref_msg += f" (e.g. {best_item.get('sku')} in {best_item.get('collection_name')})"

            return {
                "sku": f"{category}_AVG",
                "price": avg_price,
                "collection_name": col or "Unknown",
                "door_style": style or "Unknown",
                "price_ref": ref_msg
            }, "CATEGORY_AVERAGE"

    return {"sku": f"{target} (Review)", "price": 0.0, "collection_name": col or "Unknown", "price_ref": "No catalog match found"}, "MANUAL_PRICING_REQUIRED"

@router.post("/generate-bom")
async def generate_bom(project_id: str, manufacturer_id: str):
    print("DEBUG: generate_bom started")
    try:
        supabase = get_supabase()
        project = supabase.table("quotation_projects").select("*").eq("id", project_id).single().execute().data
        if not project: return {"success": False, "error": "Project not found"}
        rooms = project.get("extracted_data", {}).get("rooms", [])

        # 1. Collect required collections and styles (for room filter)
        required_cols = {'UNIVERSAL'}
        required_styles = {'UNIVERSAL'}
        project_skus = set()
        
        for room in rooms:
            room_col = str(room.get('collection') or '').upper().strip()
            if room_col:
                required_cols.add(room_col)
                # Split combined collection names (e.g. "PRIME / PREMIUM") to ensure targeted fetch catches components
                for part in re.split(r'[/\-|]', room_col):
                    p = part.strip()
                    if p and len(p) > 2:
                        required_cols.add(p)
            
            # For Integrity, we also need to fetch the combined Series-Wood collection
            wood = str(room.get('wood_species') or '').upper().strip()
            if wood and room_col and " - " not in room_col:
                required_cols.add(f"{room_col} - {wood}")
                
            if room.get('door_style'): required_styles.add(str(room['door_style']).upper().strip())
            
            # 2. Collect all SKUs from any relevant room categories
            for cat in ['cabinets', 'perimeter', 'island', 'hardware', 'island_hardware', 'bump', 'island_bump', 'opt_crown', 'opt_light_rail', 'vent_chase_material']:
                for item in room.get(cat, []):
                    code = str(item.get('code') or '').strip().upper()
                    if code:
                        project_skus.add(code)
                        # Add common variants (clean parens, stripped suffixes) for broad fetch
                        clean = re.sub(r'[\(\{\[].*?[\)\}\]]', '', code).strip()
                        if clean:
                            project_skus.add(clean)
                        no_suffix = re.sub(r'\s*(BUTT|H|L|R|FL|S|D)$', '', clean).strip()
                        if no_suffix:
                            project_skus.add(no_suffix)
                        
                        # Add spaced variants for fetch (e.g. SB36BUTT -> SB36 BUTT)
                        if "BUTT" in clean and " " not in clean:
                            project_skus.add(clean.replace("BUTT", " BUTT"))
                        if clean.startswith("SB") and not clean.startswith("SBN"):
                            project_skus.add(clean.replace("SB", "SBN", 1))

        print(f"DEBUG: Project SKUs (Targeted): {len(project_skus)}")
        print(f"DEBUG: Project Collections: {required_cols}")

        # Determine manufacturer hint (for Integrity-specific logic)
        project_mfg_name = str(project.get('metadata', {}).get('mfg_name', '')).lower()
        is_integrity = (
            manufacturer_id == "4c5206ec-fd02-46cb-81bf-f4663a7333d0"
            or "integrity" in project_mfg_name
        )
        mfg_hint = "integrity" if is_integrity else project_mfg_name

        # 3. GET PRICING MAPS (with CACHE & Targeted optimization)
        start_t = datetime.datetime.now()
        lookup_maps = get_manufacturer_pricing_maps(supabase, manufacturer_id, mfg_hint, project_skus, required_cols)
        print(f"DEBUG: Lookup maps ready in {(datetime.datetime.now() - start_t).total_seconds():.2f}s")
        
        # 4. Generate BOM Items using optimized helper
        bom_items = _price_rooms(rooms, lookup_maps, manufacturer_id, mfg_hint, project_id)
        
        print(f"DEBUG: BOM matching finished: {len(bom_items)} items")
        import json
        with open("bom_items_debug.json", "w") as f:
            json.dump(bom_items, f, indent=2)

        supabase.table("quotation_boms").delete().eq("project_id", project_id).execute()
        print("DEBUG: Previous BOM deleted")
        
        if bom_items:
            # Chunk insert if many items
            try:
                for i in range(0, len(bom_items), 500):
                    chunk = bom_items[i : i + 500]
                    supabase.table("quotation_boms").insert(chunk).execute()
                print("DEBUG: New BOM items inserted")
            except Exception as e:
                print(f"DEBUG: Chunk insert failed: {e}. Trying one by one...")
                for item in bom_items:
                    try:
                        supabase.table("quotation_boms").insert(item).execute()
                    except Exception as ie:
                        print(f"DEBUG: FAILED ITEM: {item}")
                        raise ie
            
        return {"success": True, "count": len(bom_items)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"DEBUG ERROR: {e}")
        return {"success": False, "error": str(e)}








def _quick_cabinet_type(sku: str) -> str:
    from app.utils.cabinet_utils import classify_cabinet_type
    ct = classify_cabinet_type(sku)
    if ct: return ct
    
    s = sku.upper().strip()
    if re.match(r'^W[BC]', s) or re.match(r'^WAC', s): return "Wall Corner Cabinet"
    if re.match(r'^W\d', s): return "Wall Cabinet"
    if s.startswith(("VSB", "VB", "V3S")): return "Vanity Cabinet"
    if s.startswith("SB"): return "Sink Base Cabinet"
    if re.match(r'^BB', s) or re.match(r'^LBC', s) or s.startswith("LS"): return "Base Corner Cabinet"
    if re.match(r'^B\d', s): return "Base Cabinet"
    if s.startswith(("OVD", "OV")): return "Oven Cabinet"
    if re.match(r'^(UT|UTC|T|P)\d', s): return "Tall Cabinet"
    if s.startswith(("UF", "WF", "BF", "TF", "VF")): return "Filler"
    if s.startswith(("BTK", "TK")): return "Toe Kick"
    if s.startswith("OCM"): return "Outside Corner Molding"
    if s.startswith("ICM"): return "Inside Corner Molding"
    if s.startswith(("SCM", "SHM", "SM")): return "Scribe Molding"
    if s.startswith("QM"): return "Quarter Round Molding"
    if s.startswith("LR") or "LIGHTRAIL" in s or "LIGHT RAIL" in s: return "Light Rail"
    if "CROWN" in s or s.startswith("CM"): return "Crown Molding"
    if s.startswith("HWC"): return "Hardwood Cleat"
    if s.startswith(("WTEP", "TEP", "EP")): return "End Panel"
    if s.startswith(("BACKB", "BACK-B", "BACKF", "FBP")): return "Back Panel"
    if s.startswith("SHELF"): return "Adjustable Shelf"
    if s in ("DOORS", "DOOR"): return "Door Count"
    if s in ("DRAWERS", "DRAWER", "DWR"): return "Drawer Count"
    if s.startswith("REF"): return "Refrigerator Cabinet"
    if "RANGE" in s or "HOOD" in s: return "Appliance Cabinet"
    if s.startswith("TOUCH"): return "Touch-Up Kit"
    if s.startswith("FL"): return "Filler Light Rail"
    return "Cabinet Accessory"


# Full cabinet intelligence: what it is, why it's there, what it does
def _classify_cabinet_full(sku: str) -> dict:
    from app.utils.cabinet_utils import classify_cabinet_type, parse_sku_dimensions
    
    s = sku.upper().strip()
    ct = classify_cabinet_type(s)
    dims = parse_sku_dimensions(s)
    
    prefix = dims.get('prefix')
    width = dims.get('width')
    height = dims.get('height')
    
    # Suffix meaning
    suffix_note = ''
    if 'BUTT' in s:
        suffix_note = 'Butt door style — doors meet flush at centerline with no overlay gap'
    elif re.search(r'MDBD', s):
        suffix_note = 'Modified door/blind door configuration'
    elif re.search(r'BLD$', s):
        suffix_note = 'Blind door configuration'
    elif re.search(r'AS(BUTT)?$', s):
        suffix_note = 'Angled / asymmetric configuration'
    elif re.search(r'FHD', s):
        suffix_note = 'Full Height Door (no drawer)'

    p = prefix or ''
    cat = ct or _quick_cabinet_type(s)
    
    # Base ctx logic
    ctx = ""
    if p.startswith('W'):
        h_ctx = (', ' + str(height) + '" tall') if height else ''
        if height and height >= 48:
            ctx = f'Extended-height wall cabinet ({height}") — used above refrigerator or runs to ceiling'
        elif height == 42:
            ctx = '42" tall wall cabinet — full-height unit, commonly used to ceiling'
        elif height == 30:
            ctx = '30" standard wall cabinet — typical upper installation'
        else:
            ctx = f'Wall cabinet{h_ctx} — mounted above countertop'
    
    elif p in ('SB', 'VSB', 'VB'):
        ctx = f'Sink base cabinet — open interior for plumbing access'
        if p.startswith('V'): ctx = 'Bathroom vanity sink base'
        
    elif p in ('B', 'DB', 'B3S', 'B4S'):
        b_w = (f' ({width}" wide)') if width else ''
        if '3S' in p or '4S' in p or p == 'DB':
            ctx = f'Drawer base cabinet{b_w} — features multiple storage drawers'
        else:
            ctx = f'Floor-mounted base cabinet{b_w} — supports countertop, standard 34.5" height'
            
    elif 'LS' in p or p in ('BC', 'BBC', 'LBC', 'WLS'):
        ctx = 'Corner unit — designed to maximize storage in 90-degree corner transitions'
        
    elif (p.startswith('F') and p != 'FBP') or p == 'UF':
        w_s = f"{width}\" " if width else ""
        ctx = f"{w_s}filler strip — closes gap between cabinet and wall or appliance"
        
    elif p in ('CM', 'LR', 'OCM', 'SCM', 'TK', 'BTK'):
        ctx = f'Decorative trim/molding — used for finishing and closing gaps'
        
    elif p in ('T', 'P', 'UTIL', 'OVD', 'REF'):
        ctx = 'Tall pantry/utility/refrigerator cabinet — full floor-to-ceiling storage column'

    return {'cat': cat, 'ctx': ctx, 'suffix': suffix_note}


def _quick_description(sku: str, room: str, _match_type: str, _matched_ref: str = '', _price: float = 0) -> str:
    """Generate a human-readable description: what the item is, its purpose, and where it's used."""
    info = _classify_cabinet_full(sku)
    cat  = info['cat']
    ctx  = info['ctx']
    sfx  = info['suffix']

    parts = [cat]
    if ctx:
        parts.append(ctx)
    if sfx:
        parts.append(sfx)
    if room:
        parts.append(f"Location: {room}")

    return ' — '.join(parts)


def _quick_match_explanation(drawing_sku: str, catalog_ref: str, match_type: str, price: float = 0) -> str:
    """Generate a clear, actionable explanation of HOW this item was priced."""
    mt = match_type.upper()

    # Parse catalog ref string formats
    clean    = re.sub(r'^(?:Ref:|Catalog Ref:)\s*', '', catalog_ref).strip()
    cat_sku  = clean.split('[')[0].split('(')[0].strip()
    col_m    = re.search(r'\[([^\]]+)\]', clean)
    col_name = col_m.group(1).strip() if col_m else ''
    avg_m    = re.search(r'avg\.\s*of\s*(\d+)', clean, re.I)
    avg_n    = avg_m.group(1) if avg_m else None

    p_str  = (' — List $' + '{:,.0f}'.format(price)) if price > 0 else ''
    col_str = (' [' + col_name + ']') if col_name else ''

    # Pre-compute quoted identifiers (backslash in f-strings unsupported in Python <3.12)
    dsku = '"' + drawing_sku + '"'
    csku = '"' + cat_sku + '"'

    if 'EXACT_SPEC_ORIGINAL' in mt or ('EXACT' in mt and 'ORIGINAL' in mt):
        return (
            f'EXACT MATCH: Drawing code {dsku} found directly in manufacturer catalog as {csku}{col_str}{p_str}. '
            f'Price is fully confirmed — no estimation.'
        )

    if 'BRIDGED' in mt or 'DRAW_SFX' in mt or 'SUFFIX_BRIDGE' in mt:
        if 'BUTT' in drawing_sku.upper():
            draw_sfx = (
                'Drawing annotates "BUTT" to indicate door style (doors flush at centerline). '
                'This is a floor plan annotation — not part of the base cabinet code. '
                'Manufacturer catalog uses their own door-style suffix (e.g., "BD") for the same item.'
            )
        elif re.search(r'MDBD|BLD', drawing_sku.upper()):
            draw_sfx = 'Drawing door-style suffix translated to manufacturer catalog format.'
        else:
            draw_sfx = 'Drawing code suffix translated to manufacturer catalog format.'
        return (
            f'SUFFIX TRANSLATION -> EXACT MATCH: {draw_sfx} '
            f'Matched as {csku}{col_str}{p_str}. Price is accurate.'
        )

    if 'EXACT' in mt or 'COMPRESSED' in mt:
        note = ' (after normalizing code format)' if 'COMPRESSED' in mt else ''
        return (
            f'EXACT MATCH{note}: {dsku} matched to manufacturer catalog as {csku}{col_str}{p_str}. '
            f'Price is confirmed — no estimation.'
        )

    if 'MFG_STRIPPED' in mt:
        return (
            f'CODE NORMALIZATION -> EXACT MATCH: Manufacturer suffix stripped from {dsku} '
            f'to find base code {csku}{col_str}{p_str}. Price is accurate.'
        )

    if 'FUZZY' in mt:
        score_m = re.search(r'(\d+)', mt)
        score   = int(score_m.group(1)) if score_m else 80
        quality = 'high-confidence' if score >= 90 else 'moderate'
        action  = '' if score >= 90 else ' Verify catalog code is correct before finalizing.'
        return (
            f'CLOSE MATCH ({score}% similarity): {dsku} matched to catalog SKU {csku}{col_str}{p_str}. '
            f'This is a {quality} approximate match based on code similarity.{action}'
        )

    if 'ALIAS' in mt:
        return (
            f'MOLDING/ALIAS MATCH: Drawing code {dsku} recognized as an alternate prefix for '
            f'catalog item {csku}{col_str}{p_str}. Different manufacturers use different prefix conventions for the same molding type.'
        )

    if 'NEAREST_DIM' in mt:
        return (
            f'NEAREST SIZE MATCH: No catalog entry found for {dsku} at this exact size. '
            f'Priced using closest available size in catalog: {csku}{col_str}{p_str}. '
            f'ACTION: Verify with manufacturer whether this exact cabinet size is available, '
            f'or confirm that the nearest size is an acceptable substitution.'
        )

    if 'DIMENSION' in mt:
        return (
            f'DIMENSION MATCH: {dsku} matched to {csku}{col_str} by cabinet dimensions{p_str}. '
            f'Cabinet type and size are consistent — price should be accurate.'
        )

    if 'CATEGORY_AVERAGE' in mt or 'CATEGORY' in mt:
        n_s = (avg_n + ' similar catalog items') if avg_n else 'multiple similar catalog items'
        return (
            f'ESTIMATED PRICE — ACTION REQUIRED: No catalog match found for {dsku}. '
            f'Price estimated from average of {n_s}{col_str}. '
            f'This price is NOT confirmed. Request exact pricing from manufacturer before using in a proposal.'
        )

    if 'MANUAL' in mt:
        # Special message for door/drawer counts — tell the user exactly what to enter
        sku_up = drawing_sku.upper()
        if sku_up in ('DOORS', 'DOOR'):
            door_style_hint = col_name or cat_sku or 'your selected door style'
            return (
                f'DOOR PRICE NOT IN CATALOG: No per-door price found for the selected finish/style '
                f'({door_style_hint}). '
                f'Enter the manufacturer\'s list price per door in the Unit Price column above. '
                f'Multiply by quantity to get total door cost.'
            )
        if sku_up in ('DRAWERS', 'DRAWER'):
            door_style_hint = col_name or cat_sku or 'your selected door style'
            return (
                f'DRAWER PRICE NOT IN CATALOG: No per-drawer price found for the selected finish/style '
                f'({door_style_hint}). '
                f'Enter the manufacturer\'s list price per drawer box in the Unit Price column above. '
                f'Multiply by quantity to get total drawer cost.'
            )
        return (
            f'MANUAL PRICING REQUIRED: {dsku} was not found in the catalog. '
            f'Set price manually after contacting manufacturer for current list pricing.'
        )

    return f'Matched {dsku} to catalog reference {csku}{col_str}{p_str}.'


def _match_type_confidence(match_type: str) -> float:
    mt = match_type.upper()
    if "EXACT_SPEC_ORIGINAL" in mt:
        return 1.0
    if "EXACT_SPEC_CLEANED" in mt:
        return 0.97
    if "EXACT_SPEC_BRIDGED" in mt or "SUFFIX_BRIDGE" in mt:
        return 0.93
    if "EXACT_STYLE_MFG" in mt:
        return 0.90
    if "COMPRESSED" in mt:
        return 0.85
    if "FUZZY" in mt:
        score = re.search(r'(\d+)', mt)
        return round(float(score.group(1)) / 100, 2) if score else 0.80
    if "NEAREST_DIM" in mt:
        return 0.70
    if "DIMENSION" in mt:
        return 0.65
    if "CATEGORY" in mt or "AVG" in mt:
        return 0.30
    return 0.50


def _price_rooms(rooms: list, lookup_maps: dict, manufacturer_id: str, mfg_hint: str, project_id: str) -> list:
    """Match every cabinet in every room against lookup_maps and return BOM item dicts."""
    bom_items = []
    categories_to_flatten = [
        'cabinets', 'perimeter', 'island', 'hardware', 'island_hardware',
        'bump', 'island_bump', 'opt_crown', 'opt_light_rail', 'vent_chase_material'
    ]

    for room in rooms:
        room_col  = room.get('collection', '')
        wood      = room.get('wood_species', '')
        effective_col = room_col
        if manufacturer_id == "4c5206ec-fd02-46cb-81bf-f4663a7333d0" or "integrity" in mfg_hint:
            if wood and room_col and " - " not in room_col:
                effective_col = f"{room_col} - {wood}"

        for cat in categories_to_flatten:
            for item in room.get(cat, []):
                match, match_type = find_best_match(
                    item['code'], effective_col, room.get('door_style', ''),
                    lookup_maps, manufacturer_hint=mfg_hint
                )
                if match:
                    qty   = int(float(item.get('quantity', item.get('qty', 1))))
                    price = round(float(match['price']), 2)
                    raw_code = item['code']
                    # Attach labor points if available
                    l_points = lookup_maps.get('labor', {}).get(match['sku'], 0)
                    if not l_points and match.get('csku'):
                        l_points = lookup_maps.get('labor', {}).get(match['csku'], 0)

                    bom_items.append({
                        "project_id":      project_id,
                        "sku":             raw_code,
                        "matched_sku":     match.get('price_ref') or match['sku'],
                        "qty":             qty,
                        "unit_price":      price,
                        "line_total":      round(price * qty, 2),
                        "install_points":  l_points,
                        "room":            room['room_name'],
                        "collection":      room.get('collection') or match.get('collection_name'),
                        "door_style":      room.get('door_style') or 'UNIVERSAL',
                        "price_source":    f"Python Engine ({match_type})",
                        "precision_level": match_type,
                        "description":     _quick_description(raw_code, room['room_name'], match_type),
                        "cabinet_type":    cat,
                        "match_explanation": _quick_match_explanation(raw_code, match.get('price_ref') or match['sku'], match_type),
                        "match_confidence": _match_type_confidence(match_type),
                        "created_at":      datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                    })
    return bom_items


@router.post("/compare-bom")
async def compare_bom(body: dict):
    """
    Side-by-side pricing comparison between two manufacturers for the same project.

    Body:
      {
        "project_id": "uuid",
        "manufacturer_a": {"id": "uuid", "collection": "PRIME CHERRY", "door_style": "FACE FRAME"},
        "manufacturer_b": {"id": "uuid", "collection": "MONTEREY",     "door_style": "FACE FRAME"}
      }

    Returns per-row delta and summary totals.
    """
    import asyncio
    try:
        supabase     = get_supabase()
        project_id   = body.get("project_id", "")
        mfr_a        = body.get("manufacturer_a", {})
        mfr_b        = body.get("manufacturer_b", {})

        if not project_id or not mfr_a.get("id") or not mfr_b.get("id"):
            return {"success": False, "error": "project_id, manufacturer_a.id and manufacturer_b.id are required"}

        project = supabase.table("quotation_projects").select("*").eq("id", project_id).single().execute().data
        if not project:
            return {"success": False, "error": "Project not found"}

        base_rooms = project.get("extracted_data", {}).get("rooms", [])

        def make_rooms(collection: str, door_style: str):
            return [
                {**r, "collection": collection, "door_style": door_style}
                for r in base_rooms
            ]

        # Collect all project SKUs for targeted fetch on large catalogs
        project_skus: set = set()
        for r in base_rooms:
            for cat in ['cabinets', 'perimeter', 'island', 'hardware', 'island_hardware',
                        'bump', 'island_bump', 'opt_crown', 'opt_light_rail', 'vent_chase_material']:
                for itm in r.get(cat, []):
                    code = str(itm.get('code') or '').strip().upper()
                    if code:
                        project_skus.add(code)
                        project_skus.add(re.sub(r'[\(\{\[].*?[\)\}\]]', '', code).strip())
                        project_skus.add(re.sub(r'\s*(BUTT|H|L|R|FL|S|D)$', '', code).strip())

        # Fetch pricing for both manufacturers in parallel
        pricing_a, pricing_b = await asyncio.gather(
            asyncio.to_thread(_fetch_pricing_data_internal, supabase, mfr_a["id"], project_skus),
            asyncio.to_thread(_fetch_pricing_data_internal, supabase, mfr_b["id"], project_skus),
        )

        maps_a = _build_lookup_maps(pricing_a)
        maps_b = _build_lookup_maps(pricing_b)

        mfg_hint_a = "integrity" if mfr_a["id"] == "4c5206ec-fd02-46cb-81bf-f4663a7333d0" else ""
        mfg_hint_b = "integrity" if mfr_b["id"] == "4c5206ec-fd02-46cb-81bf-f4663a7333d0" else ""

        rooms_a = make_rooms(mfr_a.get("collection", "UNIVERSAL"), mfr_a.get("door_style", "UNIVERSAL"))
        rooms_b = make_rooms(mfr_b.get("collection", "UNIVERSAL"), mfr_b.get("door_style", "UNIVERSAL"))

        items_a = _price_rooms(rooms_a, maps_a, mfr_a["id"], mfg_hint_a, project_id)
        items_b = _price_rooms(rooms_b, maps_b, mfr_b["id"], mfg_hint_b, project_id)

        def row_key(item):
            return f"{item['room']}|{item['sku']}|{item['qty']}"

        map_a = {row_key(i): i for i in items_a}
        map_b = {row_key(i): i for i in items_b}
        all_keys = sorted(set(list(map_a.keys()) + list(map_b.keys())))

        rows = []
        for key in all_keys:
            ia = map_a.get(key)
            ib = map_b.get(key)
            base    = ia or ib
            price_a = float(ia["unit_price"]) if ia else 0.0
            price_b = float(ib["unit_price"]) if ib else 0.0
            qty     = int(base["qty"])
            ext_a   = round(price_a * qty)
            ext_b   = round(price_b * qty)
            rows.append({
                "room":         base["room"],
                "sku_original": base["sku"],
                "qty":          qty,
                "sku_a":        ia["matched_sku"] if ia else None,
                "price_a":      price_a,
                "precision_a":  ia["precision_level"] if ia else None,
                "ext_a":        ext_a,
                "sku_b":        ib["matched_sku"] if ib else None,
                "price_b":      price_b,
                "precision_b":  ib["precision_level"] if ib else None,
                "ext_b":        ext_b,
                "delta":        ext_b - ext_a,
            })

        total_a = sum(r["ext_a"] for r in rows)
        total_b = sum(r["ext_b"] for r in rows)
        delta   = total_b - total_a

        return {
            "success": True,
            "rows":    rows,
            "totals": {
                "list_a":    total_a,
                "list_b":    total_b,
                "delta":     delta,
                "delta_pct": round(delta / total_a * 100, 2) if total_a else 0,
            },
        }
    except Exception as e:
        traceback.print_exc()
        return {"success": False, "error": str(e)}

@router.post("/upload-pricing")
async def upload_pricing(manufacturer_id: str, file: UploadFile = File(...)):
    """
    Upload and extract an Excel/CSV pricing file for a manufacturer.
    Returns a detailed extraction report including collection breakdown,
    SKU counts, price ranges, and sample data.
    """
    import tempfile
    from collections import defaultdict

    file_id = str(uuid.uuid4())
    # Use Python's tempfile for cross-platform compatibility (Linux /tmp and Windows %TEMP%)
    tmp_dir = tempfile.gettempdir()
    safe_name = re.sub(r'[^\w.\-]', '_', file.filename or "upload")
    temp_path = os.path.join(tmp_dir, f"{file_id}_{safe_name}")

    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    try:
        with open(temp_path, "rb") as f:
            file_bytes = f.read()

        supabase = get_supabase()

        # Look up manufacturer name so the parser can apply brand-specific logic
        # (e.g. Wellborn column mapping)

        mfg_name = ""
        try:
            mfg_row = supabase.table("manufacturers").select("name").eq("id", manufacturer_id).single().execute()
            mfg_name = (mfg_row.data or {}).get("name", "")
        except Exception:
            pass
        print(f"DEBUG: Manufacturer name resolved: '{mfg_name}'")

        # 1. Register file record (file_url is a placeholder — no cloud storage needed for admin use)
        supabase.table("manufacturer_files").insert({
            "id":              file_id,
            "manufacturer_id": manufacturer_id,
            "file_type":       "pricing",
            "file_name":       file.filename,
            "file_url":        "#",
            "file_format":     file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else None
        }).execute()

        file_ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ""
        print(f"DEBUG: Pricing extraction starting — {file.filename} ({file_ext})")

        if file_ext == "pdf":
            pricing = await parse_pricing_pdf(file_bytes, manufacturer_id, file_id)
        else:
            pricing = await parse_specifications_python(file_bytes, manufacturer_id, file_id, mfg_name)

        count = len(pricing)
        print(f"DEBUG: Extracted {count} pricing records.")

        # ── Build extraction report ────────────────────────────────────────────
        col_stats = defaultdict(lambda: {"skus": set(), "price_min": float("inf"), "price_max": 0.0, "rows": 0})
        style_set = set()

        for rec in pricing:
            col  = str(rec.get("collection_name") or "UNIVERSAL").strip()
            styl = str(rec.get("door_style") or "UNIVERSAL").strip()
            sku  = str(rec.get("sku") or "").strip()
            price = float(rec.get("price") or 0)

            col_stats[col]["skus"].add(sku)
            col_stats[col]["rows"] += 1
            if price > 0:
                col_stats[col]["price_min"] = min(col_stats[col]["price_min"], price)
                col_stats[col]["price_max"] = max(col_stats[col]["price_max"], price)
            style_set.add(styl)

        report_collections = []
        all_skus = set()
        for col_name, st in sorted(col_stats.items(), key=lambda x: -len(x[1]["skus"])):
            all_skus |= st["skus"]
            report_collections.append({
                "name":      col_name,
                "sku_count": len(st["skus"]),
                "rows":      st["rows"],
                "price_min": round(st["price_min"]) if st["price_min"] != float("inf") else 0,
                "price_max": round(st["price_max"]),
            })

        report = {
            "collections":  report_collections,
            "door_styles":  sorted(style_set),
            "total_skus":   len(all_skus),
            "total_rows":   count,
        }
        # ── End report ────────────────────────────────────────────────────────

        if count > 0:
            # Deduplicate: keep highest price per (sku, collection_name, door_style) combo.
            # Prevents duplicate rows when the same SKU appears in multiple sheets.
            dedup: dict[tuple, dict] = {}
            for rec in pricing:
                key = (rec['sku'], rec.get('collection_name', ''), rec.get('door_style', ''))
                existing = dedup.get(key)
                if existing is None or float(rec.get('price', 0)) > float(existing.get('price', 0)):
                    dedup[key] = rec
            pricing = list(dedup.values())
            count = len(pricing)
            print(f"DEBUG: After dedup: {count} unique pricing records")

            print(f"DEBUG: Replacing old pricing rows for {manufacturer_id}…")
            supabase.table("manufacturer_pricing").delete().eq("manufacturer_id", manufacturer_id).execute()

            # Sequential chunked insertion — parallel ThreadPoolExecutor silently swallows
            # per-thread exceptions, leaving only the first 1,000 rows saved.
            chunk_size = 500
            chunks = [pricing[i:i + chunk_size] for i in range(0, count, chunk_size)]
            inserted = 0
            failed_chunks = 0
            print(f"DEBUG: Inserting {count} rows in {len(chunks)} chunks of {chunk_size}…")
            for idx, chunk in enumerate(chunks):
                try:
                    supabase.table("manufacturer_pricing").insert(chunk).execute()
                    inserted += len(chunk)
                    if (idx + 1) % 10 == 0 or idx == len(chunks) - 1:
                        print(f"DEBUG: Inserted chunk {idx + 1}/{len(chunks)} — {inserted}/{count} rows done")
                except Exception as chunk_err:
                    failed_chunks += 1
                    print(f"ERROR: Chunk {idx + 1} failed ({len(chunk)} rows): {chunk_err}")

            if failed_chunks:
                print(f"WARNING: {failed_chunks} chunk(s) failed — {inserted}/{count} rows saved")
            else:
                print(f"DEBUG: All {count} rows inserted successfully.")

            # Bust all in-memory and disk caches so next pricing run uses fresh data
            global _CONFIG_CACHE, _MFG_DB_CACHE, GLOBAL_LOOKUP_CACHE
            _CONFIG_CACHE.pop(manufacturer_id, None)
            GLOBAL_LOOKUP_CACHE.pop(manufacturer_id, None)
            for k in [k for k in _MFG_DB_CACHE['sku'] if k.startswith(f"{manufacturer_id}:")]:
                del _MFG_DB_CACHE['sku'][k]
            for k in [k for k in _MFG_DB_CACHE['col'] if k.startswith(f"{manufacturer_id}:")]:
                del _MFG_DB_CACHE['col'][k]
            cpath = get_cache_path(manufacturer_id)
            if os.path.exists(cpath):
                try: os.remove(cpath)
                except: pass

        return {
            "success":  True,
            "count":    count,
            "fileName": file.filename,
            "report":   report,
        }

    except Exception as e:
        print(f"Pricing Upload Error: {e}")
        traceback.print_exc()
        return {"success": False, "error": str(e)}
    finally:
        if os.path.exists(temp_path):
            try: os.remove(temp_path)
            except: pass

def _process_spec_book_bg(job_id: str, file_bytes: bytes, manufacturer_id: str,
                           file_id: str):
    """Background worker: parses spec PDF and inserts records, updating job status."""
    from app.utils.pdf_processor import parse_pricing_pdf_sync
    try:
        SPEC_UPLOAD_JOBS[job_id]['status'] = 'processing'
        SPEC_UPLOAD_JOBS[job_id]['message'] = 'Parsing PDF pages…'
        SPEC_UPLOAD_JOBS[job_id]['progress'] = 5

        pricing = parse_pricing_pdf_sync(file_bytes, manufacturer_id, file_id)
        count = len(pricing)
        SPEC_UPLOAD_JOBS[job_id]['progress'] = 70
        SPEC_UPLOAD_JOBS[job_id]['message'] = f'Inserting {count} pricing records…'

        if count > 0:
            supabase = get_supabase()
            chunk_size = 500
            chunks = [pricing[i: i + chunk_size] for i in range(0, count, chunk_size)]
            inserted = 0
            for idx, chunk in enumerate(chunks):
                try:
                    supabase.table("manufacturer_pricing").insert(chunk).execute()
                    inserted += len(chunk)
                except Exception as chunk_err:
                    print(f"ERROR: Spec-book chunk {idx + 1} failed: {chunk_err}")
                pct = 70 + int((idx + 1) / len(chunks) * 25)
                SPEC_UPLOAD_JOBS[job_id]['progress'] = pct
            SPEC_UPLOAD_JOBS[job_id]['message'] = f'Inserted {inserted}/{count} pricing records.'

            # Bust caches
            global _CONFIG_CACHE, _MFG_DB_CACHE, GLOBAL_LOOKUP_CACHE
            _CONFIG_CACHE.pop(manufacturer_id, None)
            GLOBAL_LOOKUP_CACHE.pop(manufacturer_id, None)
            for k in [k for k in _MFG_DB_CACHE['sku'] if k.startswith(f"{manufacturer_id}:")]:
                del _MFG_DB_CACHE['sku'][k]
            for k in [k for k in _MFG_DB_CACHE['col'] if k.startswith(f"{manufacturer_id}:")]:
                del _MFG_DB_CACHE['col'][k]
            cpath = get_cache_path(manufacturer_id)
            if os.path.exists(cpath):
                try:
                    os.remove(cpath)
                except Exception:
                    pass

        SPEC_UPLOAD_JOBS[job_id].update({
            'status': 'done', 'progress': 100,
            'message': f'Done — {count} pricing records extracted.',
            'count': count,
        })
    except Exception as e:
        SPEC_UPLOAD_JOBS[job_id].update({
            'status': 'error', 'progress': 0,
            'message': str(e), 'error': str(e),
        })
        print(f"[spec-book bg] Error for job {job_id}: {e}")


@router.post("/upload-spec-book")
async def upload_spec_book(
    manufacturer_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    """
    Async spec-book PDF upload.  Returns a job_id immediately; caller polls
    /api/spec-job/{job_id} for progress.  File is read once here (not buffered
    through Next.js) and processed in a background thread.
    """
    file_id = str(uuid.uuid4())
    job_id = str(uuid.uuid4())

    # Read file once into memory — on Render this stays in process RAM, not
    # double-buffered through the Next.js route handler.
    file_bytes = await file.read()

    try:
        supabase = get_supabase()
        supabase.table("manufacturer_files").insert({
            "id": file_id,
            "manufacturer_id": manufacturer_id,
            "file_type": "spec",
            "file_name": file.filename,
            "file_url": "#",
            "file_format": file.filename.split('.')[-1] if '.' in file.filename else None
        }).execute()
    except Exception as e:
        print(f"[upload-spec-book] DB insert error: {e}")
        return {"success": False, "error": str(e)}

    SPEC_UPLOAD_JOBS[job_id] = {
        'status': 'queued', 'progress': 2,
        'message': 'Upload received, queued for processing…',
        'count': 0, 'fileName': file.filename, 'error': None,
        'file_id': file_id,
    }

    background_tasks.add_task(
        _process_spec_book_bg, job_id, file_bytes, manufacturer_id, file_id
    )

    return {"success": True, "job_id": job_id, "fileName": file.filename, "status": "queued"}


@router.get("/spec-job/{job_id}")
async def get_spec_job_status(job_id: str):
    """Returns current status/progress for an async spec-book upload job."""
    job = SPEC_UPLOAD_JOBS.get(job_id)
    if not job:
        return {"success": False, "error": "Job not found"}
    return {"success": True, **job}


@router.get("/manufacturer-config")
async def get_manufacturer_config(id: str):
    """
    Structured Collection -> Door Styles mapping for the frontend config page.
    Optimized: uses targeted DISTINCT-like query and aggressively caches results.
    """
    try:
        supabase = get_supabase()
        
        # Check if this is Wellborn/1951 to return the structured hierarchy
        WELLBORN_IDS = ["3be07931-596a-4fa3-8d39-8d04c36cf4bb", "4afa4d4e-ff56-4a76-851d-73541f8a58ec"]
        mfr_res = supabase.table("manufacturers").select("name").eq("id", id).single().execute()
        mfr_name = (mfr_res.data.get("name", "") if (mfr_res and mfr_res.data) else "").upper()
        
        if id in WELLBORN_IDS or "WELLBORN" in mfr_name or "1951" in mfr_name:
            print(f"DEBUG: Returning STRUCTURED Wellborn config for {id}")
            return {
                "success": True,
                "mapping": WELLBORN_STRUCTURE,
                "debug": {"source": "hardcoded_wellborn_v2", "mfr_name": mfr_name}
            }

        global _CONFIG_CACHE
        if id in _CONFIG_CACHE:
            print("DEBUG: Returning manufacturer config from IN-MEMORY cache")
            return {
                "success": True, 
                "mapping": _CONFIG_CACHE[id]["mapping"],
                "debug": _CONFIG_CACHE[id]["debug"]
            }

        cpath = get_cache_path(id)
        if os.path.exists(cpath):
            try:
                with open(cpath, "r") as f:
                    data = json.load(f)
                _CONFIG_CACHE[id] = data
                print("DEBUG: Returning manufacturer config from DISK cache")
                return {
                    "success": True, 
                    "mapping": data["mapping"],
                    "debug": data["debug"]
                }
            except Exception as e:
                print(f"DEBUG: Failed to read disk cache: {e}")

        supabase = get_supabase()
        mapping = {}
        
        # FAST PATH: Fetch only unique collection_name, door_style pairs
        # Supabase API max limit is 1000 rows by default (PostgREST), but we can page
        page_size = 2000
        off = 0
        seen_combos = set()
        max_scan = 100000 # Realistic cap for most catalogs
        
        import time
        start = time.time()
        
        consecutive_seen = 0
        while off < max_scan:
            # We select ONLY the two columns we need to minimize data transfer
            # We order by collection_name so that duplicates appear together, 
            # allowing the 'consecutive_seen' bailout to trigger much faster.
            res = supabase.table("manufacturer_pricing") \
                .select("collection_name, door_style") \
                .eq("manufacturer_id", id) \
                .order("collection_name, door_style") \
                .range(off, off + page_size - 1).execute()
            
            batch = res.data or []
            if not batch: break
            
            for row in batch:
                c_raw = str(row.get('collection_name') or '').strip().upper()
                st_raw = str(row.get('door_style') or '').strip().upper()
                
                if c_raw and st_raw:
                    combo = f"{c_raw}|{st_raw}"
                    if combo in seen_combos:
                        consecutive_seen += 1
                        continue
                    
                    consecutive_seen = 0 # Reset when we find a new one
                    seen_combos.add(combo)
                    
                    cols = [s.strip() for s in c_raw.split(',') if s.strip()]
                    styles = [s.strip() for s in st_raw.split(',') if s.strip()]
                    
                    for c in cols:
                        if c not in mapping: mapping[c] = set()
                        for s in styles:
                            mapping[c].add(s)
            
            if len(batch) < page_size: break
            # If we've seen 1000 items in a row that we've already processed, 
            # and since we are ordered by collection, we likely have all door styles for this block.
            if consecutive_seen > 1000: 
                print(f"DEBUG: Bailing early at {off} rows (found {len(seen_combos)} combos)")
                break
            off += page_size

        elapsed = time.time() - start
        print(f"DEBUG: Config fetch completed in {elapsed:.2f}s ({off} rows scanned, {len(seen_combos)} unique combos)")
        print(f"DEBUG: Raw mapping keys (all collections found): {list(mapping.keys())}")
            
        final_mapping = {}
        exclude = {'UNIVERSAL', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'COL B', 'COL C', 'COL D', 'COL E', 'COL F', 'COL G'}

        for c, styles in mapping.items():
            if c not in exclude:
                final_mapping[c] = sorted(list(styles))
            else:
                print(f"DEBUG: Excluded collection: '{c}'")

        # Also expose individual Tier+Wood component variants as selectable keys.
        # e.g. "PRIME MAPLE / PAINTED / DURAFORM" → adds "PRIME MAPLE", "PRIME PAINTED",
        # "PRIME DURAFORM" so the AI-extracted short name matches the dropdown.
        extra = {}
        for c_full, styles in final_mapping.items():
            tiers = [t for t in _TIER_KEYS if t in c_full]
            woods = [w for w in _WOOD_KEYS if w in c_full]
            for t in tiers:
                for w in woods:
                    variant = f"{t} {w}"
                    if variant not in final_mapping and variant not in extra:
                        extra[variant] = styles
        final_mapping.update(extra)
                
        result_data = {
            "mapping": final_mapping,
            "debug": {
                "scan_time_seconds": round(elapsed, 2),
                "unique_combos": len(seen_combos),
                "collections_found": len(final_mapping),
                "cached": False
            }
        }
        
        _CONFIG_CACHE[id] = result_data
        try:
            with open(cpath, "w") as f:
                json.dump(result_data, f)
        except Exception as e:
            print(f"DEBUG: Failed to write disk cache: {e}")

        return {
            "success": True, 
            "mapping": result_data["mapping"],
            "debug": result_data["debug"]
        }

        
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.get("/db-check")
async def db_check():
    """Diagnostic endpoint to verify database connectivity."""
    try:
        supabase = get_supabase()
        res = supabase.table("manufacturer_pricing").select("sku", count="exact").limit(1).execute()
        return {
            "success": True,
            "message": "Database connection verified",
            "pricing_count": res.count,
            "sample": res.data
        }
    except Exception as e:
        return {"success": False, "error": str(e), "trace": traceback.format_exc()}


@router.get("/catalog-check")
async def catalog_check(manufacturer_id: str):
    """
    Diagnostic: returns catalog row count, sample SKUs, and collection names
    for a given manufacturer_id. Use this to verify pricing data is uploaded
    under the correct manufacturer ID.
    """
    try:
        supabase = get_supabase()

        # Total count
        count_res = supabase.table("manufacturer_pricing") \
            .select("id", count="exact") \
            .eq("manufacturer_id", manufacturer_id) \
            .execute()
        total = count_res.count or 0

        if total == 0:
            # Also list all manufacturers that DO have pricing data
            all_mfgs_res = supabase.table("manufacturer_pricing") \
                .select("manufacturer_id") \
                .limit(500) \
                .execute()
            all_ids = list({r["manufacturer_id"] for r in (all_mfgs_res.data or [])})
            return {
                "success": True,
                "manufacturer_id": manufacturer_id,
                "total_rows": 0,
                "warning": "No pricing data found for this manufacturer_id.",
                "hint": "Pricing may have been uploaded under a different ID.",
                "manufacturers_with_pricing": all_ids,
            }

        # Sample SKUs and collection names
        sample_res = supabase.table("manufacturer_pricing") \
            .select("sku,price,collection_name,door_style") \
            .eq("manufacturer_id", manufacturer_id) \
            .limit(20) \
            .execute()

        collections_res = supabase.table("manufacturer_pricing") \
            .select("collection_name") \
            .eq("manufacturer_id", manufacturer_id) \
            .limit(500) \
            .execute()
        unique_collections = list({r["collection_name"] for r in (collections_res.data or [])})

        # Cache status
        cached = manufacturer_id in GLOBAL_LOOKUP_CACHE
        cache_age = None
        if cached:
            import datetime as _dt
            age = _dt.datetime.now().timestamp() - GLOBAL_LOOKUP_CACHE[manufacturer_id]["timestamp"]
            cache_age = f"{int(age)}s ago"

        return {
            "success": True,
            "manufacturer_id": manufacturer_id,
            "total_rows": total,
            "collections": sorted(unique_collections),
            "sample_skus": sample_res.data or [],
            "cache_status": "cached" if cached else "not cached",
            "cache_age": cache_age,
        }
    except Exception as e:
        return {"success": False, "error": str(e), "trace": traceback.format_exc()}


@router.post("/clear-cache")
async def clear_cache(manufacturer_id: str = None):
    """Clear the in-memory pricing cache for one or all manufacturers."""
    global GLOBAL_LOOKUP_CACHE
    if manufacturer_id:
        removed = manufacturer_id in GLOBAL_LOOKUP_CACHE
        GLOBAL_LOOKUP_CACHE.pop(manufacturer_id, None)
        return {"success": True, "cleared": manufacturer_id if removed else None, "message": "Cache cleared" if removed else "No cache entry found"}
    else:
        count = len(GLOBAL_LOOKUP_CACHE)
        GLOBAL_LOOKUP_CACHE.clear()
        return {"success": True, "cleared_all": True, "entries_removed": count}


# ──────────────────────────────────────────────────────────────────────────────
# INSTALL RULES  (manufacturer defaults + client overrides)
# ──────────────────────────────────────────────────────────────────────────────

@router.post("/install-rules/upload")
async def upload_install_rules(
    file: UploadFile = File(...),
    manufacturer_id: str = None
):
    """
    Bulk-upload install rules from an Excel file.

    Expected columns (case-insensitive):
      Item Code | Item Type | Install Factor | Include in 3PL | Count Basis

    All rows are upserted with the supplied manufacturer_id.
    """
    import io
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "error": "openpyxl not installed on server."}

    if not manufacturer_id:
        return {"success": False, "error": "manufacturer_id query param is required."}

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        header_map: dict[str, int] = {}
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for col_idx, cell_val in enumerate(header_row):
            if cell_val:
                header_map[str(cell_val).strip().lower()] = col_idx

        col_aliases = {
            "item_code":       ["item code", "item_code", "sku", "code"],
            "item_type":       ["item type", "item_type", "type"],
            "install_factor":  ["install factor", "install_factor", "factor"],
            "include_in_3pl":  ["include in 3pl", "include_in_3pl", "3pl"],
            "count_basis":     ["count basis", "count_basis", "basis"],
        }

        def find_col(field: str):
            for alias in col_aliases[field]:
                if alias in header_map:
                    return header_map[alias]
            return None

        col_code    = find_col("item_code")
        col_type    = find_col("item_type")
        col_factor  = find_col("install_factor")
        col_3pl     = find_col("include_in_3pl")
        col_basis   = find_col("count_basis")

        if col_code is None or col_factor is None:
            return {"success": False, "error": "Excel must have 'Item Code' and 'Install Factor' columns."}

        supabase = get_supabase()
        rows_upserted = 0
        rows_skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            item_code = str(row[col_code]).strip().upper() if row[col_code] else None
            if not item_code or item_code == "NONE":
                rows_skipped += 1
                continue

            try:
                factor = float(row[col_factor]) if row[col_factor] is not None else 1.0
            except (ValueError, TypeError):
                rows_skipped += 1
                continue

            item_type  = str(row[col_type]).strip().lower()  if (col_type  is not None and row[col_type])  else "cabinet"
            count_basis = str(row[col_basis]).strip().lower() if (col_basis is not None and row[col_basis]) else "quantity"

            raw_3pl = row[col_3pl] if col_3pl is not None else True
            if isinstance(raw_3pl, bool):
                include_3pl = raw_3pl
            elif isinstance(raw_3pl, str):
                include_3pl = raw_3pl.strip().upper() not in ("FALSE", "NO", "0", "F", "N")
            else:
                include_3pl = bool(raw_3pl) if raw_3pl is not None else True

            payload = {
                "manufacturer_id": manufacturer_id,
                "item_code":       item_code,
                "item_type":       item_type,
                "install_factor":  factor,
                "include_in_3pl":  include_3pl,
                "count_basis":     count_basis,
            }
            supabase.table("install_rules").upsert(
                payload, on_conflict="manufacturer_id,item_code"
            ).execute()
            rows_upserted += 1

        return {
            "success":       True,
            "rows_upserted": rows_upserted,
            "rows_skipped":  rows_skipped,
        }

    except Exception as e:
        return {"success": False, "error": str(e), "trace": traceback.format_exc()}


@router.post("/test-install-rule")
async def test_install_rule(body: dict):
    """
    Test rule lookup for a single item.

    Body: { manufacturer_id, client_name?, item_code, quantity }
    Returns the matched rule, calculated install units, and 3PL inclusion.
    """
    try:
        manufacturer_id = body.get("manufacturer_id", "")
        client_name     = (body.get("client_name") or "").strip()
        item_code       = str(body.get("item_code", "")).strip().upper()
        quantity        = float(body.get("quantity", 1))

        supabase = get_supabase()

        # 1. Check client overrides first
        override = None
        if client_name:
            res = supabase.table("install_rule_overrides") \
                .select("*") \
                .eq("manufacturer_id", manufacturer_id) \
                .eq("client_name", client_name) \
                .eq("item_code", item_code) \
                .limit(1).execute()
            if res.data:
                override = res.data[0]

        # 2. Manufacturer default
        default_rule = None
        res2 = supabase.table("install_rules") \
            .select("*") \
            .eq("manufacturer_id", manufacturer_id) \
            .eq("item_code", item_code) \
            .limit(1).execute()
        if res2.data:
            default_rule = res2.data[0]

        active_rule = override or default_rule
        if not active_rule:
            return {
                "success":      True,
                "found":        False,
                "item_code":    item_code,
                "message":      f"No rule found for item '{item_code}' with manufacturer '{manufacturer_id}'.",
            }

        install_units = quantity * float(active_rule["install_factor"])
        tpl_units     = quantity if active_rule["include_in_3pl"] else 0

        return {
            "success":       True,
            "found":         True,
            "item_code":     item_code,
            "rule_source":   "client_override" if override else "manufacturer_default",
            "rule":          active_rule,
            "quantity":      quantity,
            "install_units": install_units,
            "tpl_units":     tpl_units,
            "include_in_3pl": active_rule["include_in_3pl"],
        }

    except Exception as e:
        return {"success": False, "error": str(e), "trace": traceback.format_exc()}


